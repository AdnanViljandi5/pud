from __future__ import annotations

from io import BytesIO
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from business_tool.config import APP_CONFIG, BonusBlockConfig, ReferenceCellConfig
from business_tool.models import NormalizedBlock, ParsedRegion, ParsedWorkbook, ReferenceValue, WorkbookSource, WorkbookSummary


def _get_worksheet(workbook, sheet_name: str) -> Worksheet:
    try:
        return workbook[sheet_name]
    except KeyError as exc:
        raise ValueError(f"The workbook is missing the expected sheet '{sheet_name}'.") from exc


def load_workbook_source(
    uploaded_file: BinaryIO,
    preview_row_limit: int = 25,
    source_key: str = "primary_workbook",
    source_label: str = "Primary workbook",
) -> WorkbookSource:
    file_bytes = uploaded_file.getvalue()

    if not file_bytes:
        raise ValueError("The selected file is empty.")

    try:
        workbook = pd.ExcelFile(BytesIO(file_bytes), engine="openpyxl")
    except Exception as exc:
        raise ValueError(
            "The file could not be opened. Please upload a valid Excel workbook."
        ) from exc

    if not workbook.sheet_names:
        raise ValueError("No worksheets were found in the uploaded workbook.")

    return WorkbookSource(
        source_key=source_key,
        source_label=source_label,
        file_name=getattr(uploaded_file, "name", "uploaded_workbook.xlsx"),
        file_bytes=file_bytes,
        sheet_names=list(workbook.sheet_names),
        preview_row_limit=preview_row_limit,
    )


def load_sheet_preview(workbook_source: WorkbookSource, sheet_name: str) -> pd.DataFrame:
    if sheet_name not in workbook_source.sheet_names:
        raise ValueError("The selected sheet is not available in the uploaded workbook.")

    try:
        preview_frame = pd.read_excel(
            BytesIO(workbook_source.file_bytes),
            sheet_name=sheet_name,
            nrows=workbook_source.preview_row_limit,
            engine="openpyxl",
        )
    except Exception as exc:
        raise ValueError(
            "The selected sheet could not be read. Please review the workbook and try again."
        ) from exc

    preview_frame = preview_frame.dropna(how="all").reset_index(drop=True)
    preview_frame.columns = [str(column).strip() for column in preview_frame.columns]
    return preview_frame


def _load_openpyxl_workbook(file_bytes: bytes):
    try:
        return load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(
            "The workbook could not be parsed. Please check the file and try again."
        ) from exc


def _normalize_value(value: object) -> object:
    if isinstance(value, str):
        stripped_value = value.strip()
        return stripped_value if stripped_value else None
    return value


def _make_unique_headers(headers: tuple[str, ...]) -> list[str]:
    header_counts: dict[str, int] = {}
    unique_headers: list[str] = []
    for header in headers:
        current_count = header_counts.get(header, 0) + 1
        header_counts[header] = current_count
        if current_count == 1:
            unique_headers.append(header)
        else:
            unique_headers.append(f"{header} ({current_count})")
    return unique_headers


def _resolve_block_end_row(
    worksheet: Worksheet,
    block_config: BonusBlockConfig,
    next_header_row: int | None,
) -> int:
    if block_config.data_end_row is not None:
        return block_config.data_end_row
    if next_header_row is not None:
        return next_header_row - 1
    return min(worksheet.max_row, block_config.header_row + block_config.max_data_rows)


def _read_header_cells(worksheet: Worksheet, block_config: BonusBlockConfig) -> list[str]:
    start_column_index = column_index_from_string(block_config.start_column)
    header_values: list[str] = []
    for offset in range(len(block_config.expected_columns)):
        cell_value = worksheet.cell(
            row=block_config.header_row,
            column=start_column_index + offset,
        ).value
        normalized_value = _normalize_value(cell_value)
        header_values.append("" if normalized_value is None else str(normalized_value))
    return header_values


def _parse_bonus_block(
    worksheet: Worksheet,
    block_config: BonusBlockConfig,
    next_header_row: int | None,
) -> tuple[ParsedRegion, NormalizedBlock, list[str]]:
    start_column_index = column_index_from_string(block_config.start_column)
    unique_headers = _make_unique_headers(block_config.expected_columns)
    detected_headers = _read_header_cells(worksheet, block_config)
    block_issues: list[str] = []
    missing_expected_columns: list[str] = []

    for expected_header, detected_header in zip(block_config.expected_columns, detected_headers):
        if detected_header and detected_header != expected_header:
            block_issues.append(
                f"{block_config.name}: header '{detected_header}' was found where '{expected_header}' was expected."
            )
        if not detected_header:
            block_issues.append(
                f"{block_config.name}: header row appears incomplete and may need confirmation."
            )
            missing_expected_columns.append(expected_header)
            continue
        if detected_header != expected_header:
            missing_expected_columns.append(expected_header)

    if not any(detected_headers):
        block_issues.append(
            f"{block_config.name}: the expected header row was not found in the configured position."
        )

    detected_header_counts: dict[str, int] = {}
    for header in detected_headers:
        if not header:
            continue
        detected_header_counts[header] = detected_header_counts.get(header, 0) + 1

    expected_header_counts: dict[str, int] = {}
    for header in block_config.expected_columns:
        expected_header_counts[header] = expected_header_counts.get(header, 0) + 1

    duplicate_detected_columns = [
        header_name
        for header_name, header_count in detected_header_counts.items()
        if header_count > 1
    ]
    ambiguous_columns = [
        header_name
        for header_name, header_count in expected_header_counts.items()
        if header_count > 1
    ]

    data_end_row = _resolve_block_end_row(worksheet, block_config, next_header_row)
    row_records: list[list[object]] = []
    source_row_numbers: list[int] = []

    for row_index in range(block_config.header_row + 1, data_end_row + 1):
        row_values: list[object] = []
        for offset in range(len(block_config.expected_columns)):
            cell_value = worksheet.cell(row=row_index, column=start_column_index + offset).value
            row_values.append(_normalize_value(cell_value))
        if any(value is not None for value in row_values):
            row_records.append(row_values)
            source_row_numbers.append(row_index)

    raw_frame = pd.DataFrame(row_records, columns=unique_headers)
    raw_frame.insert(0, "Source Block", block_config.name)
    raw_frame.insert(1, "Source Sheet", block_config.sheet_name)
    raw_frame.insert(2, "Header Row", block_config.header_row)
    raw_frame.insert(3, "Source Row", source_row_numbers)

    normalized_frame = pd.DataFrame(row_records, columns=block_config.internal_columns)
    normalized_frame.insert(0, "source_row_number", source_row_numbers)
    normalized_frame.insert(0, "header_row", block_config.header_row)
    normalized_frame.insert(0, "source_sheet", block_config.sheet_name)
    normalized_frame.insert(0, "block_name", block_config.name)
    normalized_frame.insert(0, "block_key", block_config.key)

    specific_fields = tuple(
        field_name for field_name in block_config.internal_columns if field_name not in block_config.shared_fields
    )
    field_labels = {
        internal_name: source_name
        for source_name, internal_name in zip(block_config.expected_columns, block_config.internal_columns)
    }

    detected_columns = unique_headers.copy()
    return (
        ParsedRegion(
            region_name=block_config.name,
            frame=raw_frame,
            detected_columns=detected_headers,
            expected_columns=list(block_config.expected_columns),
            missing_expected_columns=missing_expected_columns,
            duplicate_detected_columns=duplicate_detected_columns,
            ambiguous_columns=ambiguous_columns,
            source_sheet=block_config.sheet_name,
            header_row=block_config.header_row,
        ),
        NormalizedBlock(
            block_key=block_config.key,
            block_name=block_config.name,
            normalized_frame=normalized_frame,
            shared_fields=block_config.shared_fields,
            specific_fields=specific_fields,
            field_labels=field_labels,
            source_sheet=block_config.sheet_name,
            header_row=block_config.header_row,
        ),
        block_issues,
    )


def parse_top_reference_area(workbook, reference_cells: tuple[ReferenceCellConfig, ...]) -> list[ReferenceValue]:
    reference_values: list[ReferenceValue] = []
    for cell_config in reference_cells:
        worksheet = _get_worksheet(workbook, cell_config.sheet_name)
        value = _normalize_value(worksheet[cell_config.cell_ref].value)
        if value is not None:
            reference_values.append(
                ReferenceValue(
                    label=cell_config.label,
                    cell_ref=cell_config.cell_ref,
                    value=value,
                )
            )
    return reference_values


def parse_bonus_block_1(workbook) -> tuple[ParsedRegion, NormalizedBlock, list[str]]:
    worksheet = _get_worksheet(workbook, APP_CONFIG.bonus_blocks[0].sheet_name)
    next_header_row = APP_CONFIG.bonus_blocks[1].header_row if len(APP_CONFIG.bonus_blocks) > 1 else None
    return _parse_bonus_block(worksheet, APP_CONFIG.bonus_blocks[0], next_header_row)


def parse_bonus_block_2(workbook) -> tuple[ParsedRegion, NormalizedBlock, list[str]]:
    worksheet = _get_worksheet(workbook, APP_CONFIG.bonus_blocks[1].sheet_name)
    next_header_row = APP_CONFIG.bonus_blocks[2].header_row if len(APP_CONFIG.bonus_blocks) > 2 else None
    return _parse_bonus_block(worksheet, APP_CONFIG.bonus_blocks[1], next_header_row)


def parse_bonus_block_3(workbook) -> tuple[ParsedRegion, NormalizedBlock, list[str]]:
    worksheet = _get_worksheet(workbook, APP_CONFIG.bonus_blocks[2].sheet_name)
    next_header_row = APP_CONFIG.bonus_blocks[3].header_row if len(APP_CONFIG.bonus_blocks) > 3 else None
    return _parse_bonus_block(worksheet, APP_CONFIG.bonus_blocks[2], next_header_row)


def parse_bonus_block_4(workbook) -> tuple[ParsedRegion, NormalizedBlock, list[str]]:
    worksheet = _get_worksheet(workbook, APP_CONFIG.bonus_blocks[3].sheet_name)
    return _parse_bonus_block(worksheet, APP_CONFIG.bonus_blocks[3], next_header_row=None)


def _build_review_frame(regions: list[ParsedRegion]) -> pd.DataFrame:
    if not regions:
        return pd.DataFrame()
    combined_frame = pd.concat([region.frame for region in regions], ignore_index=True, sort=False)
    ordered_columns = ["Source Block", "Source Sheet", "Header Row"]
    remaining_columns = [column for column in combined_frame.columns if column not in ordered_columns]
    return combined_frame[ordered_columns + remaining_columns]


def _build_normalized_review_frame(normalized_blocks: list[NormalizedBlock]) -> pd.DataFrame:
    if not normalized_blocks:
        return pd.DataFrame()
    return pd.concat(
        [normalized_block.normalized_frame for normalized_block in normalized_blocks],
        ignore_index=True,
        sort=False,
    )


def parse_workbook(workbook_source: WorkbookSource) -> ParsedWorkbook:
    workbook = _load_openpyxl_workbook(workbook_source.file_bytes)
    issues: list[str] = []

    missing_sheets = {
        config.sheet_name
        for config in (*APP_CONFIG.reference_cells, *APP_CONFIG.bonus_blocks)
        if config.sheet_name not in workbook.sheetnames
    }
    if missing_sheets:
        missing_sheet_names = ", ".join(sorted(missing_sheets))
        sheet_label = "sheet" if len(missing_sheets) == 1 else "sheets"
        raise ValueError(f"The workbook is missing the expected {sheet_label}: {missing_sheet_names}.")

    reference_values = parse_top_reference_area(workbook, APP_CONFIG.reference_cells)

    block_parsers = (
        parse_bonus_block_1,
        parse_bonus_block_2,
        parse_bonus_block_3,
        parse_bonus_block_4,
    )
    parsed_regions: list[ParsedRegion] = []
    normalized_blocks: list[NormalizedBlock] = []

    for parser in block_parsers:
        parsed_region, normalized_block, block_issues = parser(workbook)
        parsed_regions.append(parsed_region)
        normalized_blocks.append(normalized_block)
        issues.extend(block_issues)

    combined_frame = (
        pd.concat([region.frame for region in parsed_regions], ignore_index=True, sort=False)
        if parsed_regions
        else pd.DataFrame()
    )
    review_frame = _build_review_frame(parsed_regions)
    normalized_combined_frame = _build_normalized_review_frame(normalized_blocks)

    return ParsedWorkbook(
        reference_values=reference_values,
        regions=parsed_regions,
        normalized_blocks=normalized_blocks,
        combined_frame=combined_frame,
        review_frame=review_frame,
        normalized_combined_frame=normalized_combined_frame,
        summary=WorkbookSummary(
            sheet_count=len(workbook.sheetnames),
            region_count=len(parsed_regions),
            row_count=len(combined_frame.index),
        ),
        issues=issues,
    )

# Mapping rules may need small adjustments once the real workbook is checked,
# especially if header captions or exact row positions differ from the draft.
# Future second-source mapping should be layered above this parser so the
# current primary workbook logic stays stable and easy to maintain.
